'''
Created on 2018年5月3日

@author: user
'''


if __name__ == '__main__':
    pass


import openpyxl
 
wd = openpyxl.Workbook()
#w1 = wd.create_sheet("Test", 0)
#w2 = wd.create_sheet("Test2")

for sheet in wd:
    print (sheet.title)



    
wd.save('balances.xlsx')

